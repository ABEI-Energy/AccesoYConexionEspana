import datetime as dt
import locale as lc
import os
from io import BytesIO
from math import *

import docx2pdf as d2p
import pandas as pd
import unidecode as ud
from docx import Document
from docx.shared import Cm
from openpyxl import load_workbook
import PyPDF2 as rpdf


import numToLet as ntl

#Set the language for datetime
lc.setlocale(lc.LC_ALL,'es_ES.UTF-8')
month = dt.datetime.now().strftime("%B %Y")

def round_school(x):
    i, f = divmod(x, 1)
    return int(i + ((f >= 0.5) if (x > 0) else (f > 0.5)))

def normalize(string):
    return str(round(float(string.replace(",", ".")),2))

def excelReaderPVD(excel, dicParamPV5, rootEstructuras, dfModulos, user):

    df = pd.read_excel(excel)
    df = df.dropna(how = 'all', axis = 'columns')
    df = df.dropna(how = 'all', axis = 'rows')
    df.drop(df.columns[0], axis = 1, inplace = True)
    df = df.dropna(how = 'all', axis = 'rows')
    df1, df2 = df.iloc[:,0:2], df.iloc[:,2:]
    df1.columns = df2.columns
    df = pd.concat([df1, df2], axis = 0, ignore_index = True) 
    df = df.set_axis(['Variable', 'Valor'], axis = 1)
    df = df.dropna(how = 'all', axis = 'rows')
    df['Categoría'] = "Categoría"

    for rows in range(len(df)):
        if ((df.iloc[rows][0].isupper()) and (not isinstance(df.iloc[rows][1],str))):
            pvdCat =  df.iloc[rows][0]
        df['Categoría'][rows] = pvdCat
    df.dropna(inplace = True)

    df = df[['Categoría', 'Variable', 'Valor']]


    keywordsPVD = ('potPico', 'nombreProyecto','moduloManuf', 'modeloModulo', 'tecnologiaModulo', 'moduloPpico', 'UMaxModulo', 'inverterManuf', 'inverterModel', 'PnNonL', 'rangoUNonL', 'UMaxNonL', 'numModulos', 'numStrings', 'numEstructuras', 'numInverter', 'numTrafos', 'estructuraTipo', 'pitchEstructura', 'ratioTrafoSET', 'potPOI')
    wordsDocPVD = dict.fromkeys(keywordsPVD)

    #Proyecto
    wordsDocPVD['nombreProyecto'] = df.loc[df['Variable'] == "Nombre", 'Valor'].item()
    wordsDocPVD['potPico'] = df.loc[df['Variable'] == "Potencia pico", 'Valor'].item()
    wordsDocPVD['potPico'] = str(float(wordsDocPVD['potPico'].split(' ')[0])/1000) 
    wordsDocPVD['potPOI'] = str(round(float(wordsDocPVD['potPico'])/1.35,2))

    
    if dicParamPV5['potProj5'] == 'Sí':
        wordsDocPVD['potInstalada'] = dicParamPV5['potProjMWac'] #a veces no es 4.99, a veces es 4.6, la mete el usuario
    else: wordsDocPVD['potInstalada'] = str(round(float(df.loc[df['Variable']== "Potencia nominal", 'Valor'].item().partition(' ')[0])/1000,2))

    wordsDocPVD['dcAcRatio'] = str(round(float(wordsDocPVD['potPico'])/float(wordsDocPVD['potInstalada']),2)) #potPico/potInstalada (potencia en inversores o la que a vceces meten ellos a mano)

    #Módulos
    wordsDocPVD['moduloManuf'] = df.loc[(df['Variable'] == "Compañía") & (df['Categoría'] == "MÓDULO FV"), 'Valor'].item()
    wordsDocPVD['modeloModulo'] = df.loc[(df['Variable'] == "Modelo") & (df['Categoría'] == "MÓDULO FV"), 'Valor'].item()
    wordsDocPVD['numModulos'] = df.loc[df['Variable'] == "Módulos", 'Valor'].item()
    wordsDocPVD['moduloPpico'] = df.loc[(df['Variable'] == "Potencia") & (df['Categoría'] == "MÓDULO FV"), 'Valor'].item().partition(' ')[0]
    wordsDocPVD['tecnologiaModulo'] = df.loc[df['Variable'] == "Tecnología", 'Valor'].item()
    wordsDocPVD['UMaxModulo'] = df.loc[(df['Variable'] == "Tensión máxima") & (df['Categoría'] == "MÓDULO FV"), 'Valor'].item().partition(' ')[0]
    wordsDocPVD['sizeModulo'] = dfModulos.loc[(dfModulos['MODEL'] == wordsDocPVD['modeloModulo']) , 'SIZE'].item()
 
    wordsDocPVD['materialModulo'] = wordsDocPVD['tecnologiaModulo'].split(',')[0]
    wordsDocPVD['tipoModulo'] = wordsDocPVD['tecnologiaModulo'].split(',')[1].strip()
    wordsDocPVD['numStrings'] = df.loc[df['Variable'] == "Strings", 'Valor'].item()
    wordsDocPVD['numModuloString'] = str(int(round_school(wordsDocPVD['numModulos']/wordsDocPVD['numStrings'])))

    #Inversores
    wordsDocPVD['inverterManuf'] = df.loc[(df['Variable'] == "Compañía") & (df['Categoría'] == "INVERSOR"), 'Valor'].item()
    wordsDocPVD['numInverter'] = df.loc[df['Variable'] == "Inversor", 'Valor'].item()
    wordsDocPVD['inverterModel'] = df.loc[(df['Variable'] == "Modelo") & (df['Categoría'] == "INVERSOR"), 'Valor'].item()
    wordsDocPVD['PnNonL'] = df.loc[(df['Variable'] == "Potencia") & (df['Categoría'] == "INVERSOR"), 'Valor'].item().partition(' ')[0]
    wordsDocPVD['rangoUNonL'] = df.loc[df['Variable'] == "Rango MPPT", 'Valor'].item().replace(' V',"")
    wordsDocPVD['UMaxNonL'] = df.loc[(df['Variable'] == "Tensión máxima") & (df['Categoría'] == "INVERSOR"), 'Valor'].item().partition(' ')[0]


    #Estructuras
    wordsDocPVD['pitchEstructura'] = df.loc[df['Variable'] == "Pitch", 'Valor'].item().partition(' ')[0] + " m"
    wordsDocPVD['numEstructuras'] = df.loc[df['Variable'] == "Estructuras", 'Valor'].item().replace('(',"").replace(')',"").partition(' ')[0] 
    wordsDocPVD['estructuraTipo'] = df.loc[df['Variable'] == "Tipo", 'Valor'].item()
    wordsDocPVD['modulosEstructura'] = str(round_school(int(wordsDocPVD['numModulos'])/int(wordsDocPVD['numStrings'])))
    wordsDocPVD['figuraStruct'] = rootEstructuras + "/" + wordsDocPVD['estructuraTipo'] + ".png"

    # Este dato debería salir de alguna datasheet, pero de momento lo metemos a cholón
    if wordsDocPVD['estructuraTipo'] == '1V':
        wordsDocPVD['longFilaTracker'] = '40.6 m'
    else:
        wordsDocPVD['longFilaTracker'] = '38.2 m'


    

    wordsDocPVD['ratioTrafoSET'] = df.loc[df['Variable'] == "Ratio transf.", 'Valor'].item().replace('kV',"")
    # Resto de info se mete a mano por el redactor


    wordsDocPVD['numInverter'] = str(wordsDocPVD['numInverter'])  

    # Estilos especiales
    #Cover Bold
    wordsDocPVD['potPicoC'] = wordsDocPVD['potPico']
    wordsDocPVD['potInstaladaC'] = wordsDocPVD['potInstalada']
    wordsDocPVD['nombreProyectoC'] = wordsDocPVD['nombreProyecto']


    #Cover Light

    wordsDocPVD['dateCoverC'] = month.capitalize()


    # Fechas
    wordsDocPVD['dateMY'] = month
    wordsDocPVD['dateCoverC'] = month.capitalize()
    wordsDocPVD['date'] = dt.datetime.now().strftime("%d/%m/%y")


    wordsDocPVD['elaboradoDoc'] = user

    print("Documentos PVD done") 

    return wordsDocPVD

def excelReaderCoordenadas(excel, mainDic):

    workbook = load_workbook(excel, data_only=True)
    vallado = workbook["Vallado"]
    if mainDic['lineaTipo'] == "Aéreo":
        try:
            linea = workbook["Tramo Aéreo"]
        except Exception as e:
            linea = workbook["Tramo Soterrado"]
    else: 
        try:
            linea = workbook["Tramo Soterrado"]
        except Exception as e:
            linea = workbook["Tramo Aéreo"]

    varios = workbook["Otros"]

    dfVallado = pd.DataFrame(vallado.values)

    dfVallado = dfVallado.dropna(how = 'all', axis = 'columns')
    dfVallado = dfVallado.dropna(how = 'all', axis = 'rows')
    dfVallado.reset_index(inplace = True, drop = True)
    dfVallado.drop([0,1], axis = 0, inplace = True)
    dfVallado.reset_index(inplace = True, drop = True)
    if len(dfVallado.columns) > 3:
        dfVallado1 = dfVallado.iloc[:,0:3]
        dfVallado2 = dfVallado.iloc[:,3:]
        dfVallado2.drop([0], axis = 0, inplace = True)
        dfVallado1.columns = dfVallado2.columns
        dfVallado = pd.concat([dfVallado1, dfVallado2], axis = 0, ignore_index = True)
        dfVallado.columns = dfVallado.iloc[0]
        dfVallado.drop([0], axis = 0, inplace = True)
    else:
        dfVallado.columns = dfVallado.iloc[0]
        dfVallado.drop([0], axis = 0, inplace = True)
    dfVallado = dfVallado.dropna(how = 'any', axis = 'rows')
    dfVallado.reset_index(inplace = True, drop = True)



    dfLinea = pd.DataFrame(linea.values) #apoyos Inicio-fin

    dfLinea = dfLinea.dropna(how = 'all', axis = 'rows')
    dfLinea.reset_index(inplace = True, drop = True)
    dfLinea.drop([0,1,2], axis = 0, inplace = True)
    dfLinea.reset_index(inplace = True, drop = True)
    dfLinea = dfLinea.dropna(how = 'all', axis = 'columns')
    dfLinea.reset_index(inplace = True, drop = True)
    dfLinea.columns = dfLinea.iloc[0]
    dfLinea.drop([0], axis = 0, inplace = True)
    
    dfAAC = dfLinea.iloc[:,0:3] #Esta es para AAC A FUTUROS
    dfLinIF = dfLinea.iloc[:,3:6] #Esta es para AyC
    dfLinIF = dfLinIF.dropna(how = 'all', axis = 'rows')
    dfLinIF.reset_index(inplace = True, drop = True)


    dfCentroTrafo = pd.DataFrame(varios.values) #Coordenadas centros planta

    dfCentroTrafo = dfCentroTrafo.dropna(how = 'all', axis = 'rows')
    dfCentroTrafo.reset_index(inplace = True, drop = True)
    dfCentroTrafo.drop([0,1,2], axis = 0, inplace = True)
    dfCentroTrafo.reset_index(inplace = True, drop = True)
    dfCentroTrafo = dfCentroTrafo.dropna(how = 'all', axis = 'columns')
    dfCentroTrafo.reset_index(inplace = True, drop = True)
    dfCentroTrafo.columns = dfCentroTrafo.iloc[0]

    dfCT = dfCentroTrafo.iloc[:,3:6] #Centro de trafo
    dfCT = dfCT.dropna(how = 'all', axis = 'rows')
    dfCT.reset_index(inplace = True, drop = True)
    dfCT.columns = dfCT.iloc[0]
    dfCT.drop([0], axis = 0, inplace = True)


    dfAcceso = pd.DataFrame(varios.values) #Coordenadas punto de acceso

    dfAcceso = dfAcceso.dropna(how = 'all', axis = 'rows')
    dfAcceso.reset_index(inplace = True, drop = True)
    dfAcceso.drop([0,1,2], axis = 0, inplace = True)
    dfAcceso.reset_index(inplace = True, drop = True)
    dfAcceso = dfAcceso.dropna(how = 'all', axis = 'columns')
    dfAcceso.reset_index(inplace = True, drop = True)
    dfAcceso.columns = dfAcceso.iloc[0]

    dfAC = dfAcceso.iloc[:,0:3] #Centro de trafo
    dfAC = dfAC.dropna(how = 'all', axis = 'rows')
    dfAC.reset_index(inplace = True, drop = True)
    dfAC.columns = dfAC.iloc[0]
    dfAC.drop([0], axis = 0, inplace = True)

    print("Documentos coordenadas done") 

    return dfVallado, dfLinIF, dfCT, dfAC

def excelReaderParcelas(excel, mainDic):

    workbook = load_workbook(excel)
    planta = workbook["Planta"]
    
    if mainDic['lineaTipo'] == "Aéreo":
        try:
            tramo = workbook["Tramo Aéreo"]
        except Exception as e:
            tramo = workbook["Tramo Soterrado"]
    else: 
        try:
            tramo = workbook["Tramo Soterrado"]
        except Exception as e:
            tramo = workbook["Tramo Aéreo"]


    dfPlanta = pd.DataFrame(planta.values)

    dfPlanta = dfPlanta.dropna(how = 'all', axis = 'columns')
    dfPlanta = dfPlanta.dropna(how = 'all', axis = 'rows')
    dfPlanta.drop([1,2], axis = 0, inplace = True)
    dfPlanta.reset_index(inplace = True, drop = True)
    dfPlanta.columns = dfPlanta.iloc[0]
    dfPlanta.drop([0], axis = 0, inplace = True)
    dfPlanta = dfPlanta[dfPlanta.columns.dropna()] #Nos cargamos las columnas que no tengan Header. df.columns returns index, dropna quita los que no son, df[index] devuelve el dataframe de esas columnas
    dfPlanta = dfPlanta.iloc[:,0:6]
    dfPlanta.reset_index(inplace = True, drop = True)
    # if "enlace" or "Enlace" in dfPlanta.columns:
    #     dfPlanta.drop(columns = dfPlanta.columns[-1], axis = 1, inplace = True)
    dfPlanta = dfPlanta.dropna(how = 'any', axis = 'rows')
    dfPlanta.reset_index(inplace = True, drop = True)


    dfTramo = pd.DataFrame(tramo.values)

    dfTramo = dfTramo.dropna(how = 'all', axis = 'columns')
    dfTramo = dfTramo.dropna(how = 'all', axis = 'rows')
    dfTramo.drop([1,2], axis = 0, inplace = True)
    dfTramo.reset_index(inplace = True, drop = True)
    dfTramo.columns = dfTramo.iloc[0]
    dfTramo.drop([0], axis = 0, inplace = True)
    dfTramo = dfTramo[dfTramo.columns.dropna()] #Nos cargamos las columnas que no tengan Header. df.columns returns index, dropna quita los que no son, df[index] devuelve el dataframe de esas columnas
    dfTramo = dfTramo.iloc[:,0:6]
    dfTramo.reset_index(inplace = True, drop = True)
    # if "enlace" or "Enlace" in dfTramo.columns:
    #     dfTramo.drop(columns = dfTramo.columns[-1], axis = 1, inplace = True)
    dfTramo = dfTramo.dropna(how = 'any', axis = 'rows')
    dfTramo.reset_index(inplace = True, drop = True)

    print("Documentos parcelas done") 

    return dfPlanta, dfTramo



def reader(pdfFile,page):

    docPdf = rpdf.PdfReader(pdfFile).pages[page].extract_text()

    return docPdf

